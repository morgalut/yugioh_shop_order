# app.py
# Run:
#   pip install fastapi uvicorn openpyxl sqlalchemy psycopg[binary]
#   uvicorn app:app --reload
# Open:
#   http://127.0.0.1:8000

from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import List, Optional, Dict, Tuple
from contextlib import asynccontextmanager

from fastapi import FastAPI, Form, Query
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi import Response
import time


from sqlalchemy import (
    create_engine,
    event,
    MetaData,
    Table,
    Column,
    Integer,
    String,
    ForeignKey,
    Text,
    select,
    insert,
    update,
    delete,
    text as sql_text,
)
from sqlalchemy.engine import Engine


# ----------------------------
# Render-friendly database config
# ----------------------------
def _normalize_database_url(url: str) -> str:
    """
    Render/Heroku sometimes use 'postgres://'. SQLAlchemy expects 'postgresql://'.
    """
    url = (url or "").strip()
    if url.startswith("postgres://"):
        url = "postgresql://" + url[len("postgres://") :]
    return url


def _default_sqlite_path() -> str:
    """
    Render Persistent Disk docs recommend mounting e.g. /var/data.
    Only filesystem changes under the disk mount path persist. :contentReference[oaicite:5]{index=5}
    """
    # If you mount a disk at /var/data, keep SQLite there.
    # Locally (no disk), this still works if /var/data exists; otherwise we fall back.
    return os.getenv("DB_PATH", "/var/data/customers.db")


def make_engine() -> Engine:
    db_url = os.getenv("DATABASE_URL", "").strip()
    if db_url:
        db_url = _normalize_database_url(db_url)
        # Prefer psycopg (modern). If user provides plain postgresql:// it still works.
        if db_url.startswith("postgresql://"):
            db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)
        engine = create_engine(db_url, pool_pre_ping=True)
        return engine

    # Fallback: SQLite (best only with Persistent Disk on Render)
    sqlite_path = _default_sqlite_path()
    sqlite_dir = os.path.dirname(sqlite_path)
    if sqlite_dir and not os.path.exists(sqlite_dir):
        # If running locally and /var/data doesn't exist, fall back to project file.
        # (On Render with a mounted disk, /var/data exists.)
        sqlite_path = "customers.db"
    else:
        # Ensure directory exists if it is a real path
        if sqlite_dir:
            os.makedirs(sqlite_dir, exist_ok=True)

    engine = create_engine(
        f"sqlite:///{sqlite_path}",
        connect_args={"check_same_thread": False},
        pool_pre_ping=True,
    )
    return engine


ENGINE = make_engine()
META = MetaData()

# Enable foreign keys on SQLite
@event.listens_for(ENGINE, "connect")
def _set_sqlite_pragma(dbapi_connection, connection_record):
    try:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys = ON;")
        cursor.close()
    except Exception:
        # Not SQLite or cannot set pragma
        pass


# ----------------------------
# Schema (works on SQLite + Postgres)
# ----------------------------
customers = Table(
    "customers",
    META,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("name", String, nullable=False, unique=True),
    Column("created_at_utc", String, nullable=False),
)

orders = Table(
    "orders",
    META,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("customer_id", Integer, ForeignKey("customers.id", ondelete="CASCADE"), nullable=False),
    Column("pasted_text", Text, nullable=False),
    Column("global_note", Text, nullable=True),
    Column("status", String, nullable=False, server_default="OPEN"),  # OPEN / ARCHIVED
    Column("created_at_utc", String, nullable=False),
)

order_lines = Table(
    "order_lines",
    META,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("order_id", Integer, ForeignKey("orders.id", ondelete="CASCADE"), nullable=False),
    Column("card_name", String, nullable=False),
    Column("qty", Integer, nullable=False),
    Column("display_item", String, nullable=False),
    Column("rarity", String, nullable=True),
    Column("notes", Text, nullable=True),
    Column("raw_line", Text, nullable=True),
    Column("created_at_utc", String, nullable=False),
)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _collapse_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def init_db() -> None:
    META.create_all(ENGINE)
    # Helpful indexes (idempotent: create if not exists is DB-specific; we’ll do best-effort)
    with ENGINE.begin() as conn:
        # SQLite supports IF NOT EXISTS; Postgres supports IF NOT EXISTS as well for CREATE INDEX.
        conn.execute(sql_text("CREATE INDEX IF NOT EXISTS idx_orders_customer_status ON orders(customer_id, status)"))
        conn.execute(sql_text("CREATE INDEX IF NOT EXISTS idx_lines_order ON order_lines(order_id)"))


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(lifespan=lifespan)


# ----------------------------
# Health endpoint for Render + your UI pings
# ----------------------------
@app.get("/ping")
def ping():
    # Basic DB check too (useful on Render health checks)
    try:
        with ENGINE.begin() as conn:
            conn.execute(sql_text("SELECT 1"))
        return {"ok": True, "utc": utc_now_iso()}
    except Exception as e:
        return {"ok": False, "error": str(e), "utc": utc_now_iso()}


# ----------------------------
# Customers
# ----------------------------
def get_customer_by_exact_name(name: str) -> Optional[Dict]:
    name = _collapse_spaces(name)
    if not name:
        return None
    with ENGINE.begin() as conn:
        row = conn.execute(
            select(customers.c.id, customers.c.name).where(customers.c.name == name)
        ).mappings().first()
        return dict(row) if row else None


def create_customer(name: str) -> Tuple[bool, Optional[Dict], str]:
    name = _collapse_spaces(name)
    if not name:
        return (False, None, "Empty name.")

    with ENGINE.begin() as conn:
        try:
            res = conn.execute(
                insert(customers).values(name=name, created_at_utc=utc_now_iso())
            )
            # Fetch inserted row
            row = conn.execute(
                select(customers.c.id, customers.c.name).where(customers.c.id == res.inserted_primary_key[0])
            ).mappings().first()
            return (True, dict(row), f'Created "{name}".')
        except Exception:
            # Most likely UNIQUE constraint
            row = get_customer_by_exact_name(name)
            return (False, row, f'"{name}" already exists.')


def search_customers(q: str, limit: int = 20) -> List[Dict]:
    q = _collapse_spaces(q)
    with ENGINE.begin() as conn:
        if not q:
            rows = conn.execute(
                select(customers.c.id, customers.c.name).order_by(customers.c.name.asc()).limit(limit)
            ).mappings().all()
            return [dict(r) for r in rows]

        like = f"%{q}%"
        rows = conn.execute(
            select(customers.c.id, customers.c.name)
            .where(customers.c.name.like(like))
            .order_by(customers.c.name.asc())
            .limit(limit)
        ).mappings().all()
        return [dict(r) for r in rows]


# ----------------------------
# Orders & lines
# ----------------------------
def create_order(customer_id: int, pasted_text: str, global_note: Optional[str]) -> int:
    with ENGINE.begin() as conn:
        res = conn.execute(
            insert(orders).values(
                customer_id=int(customer_id),
                pasted_text=pasted_text,
                global_note=global_note,
                status="OPEN",
                created_at_utc=utc_now_iso(),
            )
        )
        return int(res.inserted_primary_key[0])


def add_lines(order_id: int, lines: List["ParsedLine"]) -> None:
    now = utc_now_iso()
    payload = []
    for ln in lines:
        payload.append(
            dict(
                order_id=int(order_id),
                card_name=ln.card_name_raw,
                qty=int(ln.quantity),
                display_item=ln.display_item,
                rarity=(ln.rarity or ""),
                notes=(ln.notes or ""),
                raw_line=ln.raw_line,
                created_at_utc=now,
            )
        )
    if not payload:
        return

    with ENGINE.begin() as conn:
        conn.execute(insert(order_lines), payload)


def get_open_basket_grouped() -> List[Dict]:
    """
    Returns list of:
      {
        customer_id, customer_name,
        orders: [{order_id, created_at_utc, global_note, pasted_text}],
        lines:  [{line_id, order_id, card_name, qty, display_item, rarity, notes, raw_line}]
      }
    """
    with ENGINE.begin() as conn:
        customer_rows = conn.execute(
            sql_text(
                """
                SELECT DISTINCT c.id AS customer_id, c.name AS customer_name
                FROM customers c
                JOIN orders o ON o.customer_id = c.id
                WHERE o.status = 'OPEN'
                ORDER BY c.name ASC
                """
            )
        ).mappings().all()

        result: List[Dict] = []
        for c in customer_rows:
            cid = int(c["customer_id"])
            orders_rows = conn.execute(
                sql_text(
                    """
                    SELECT id AS order_id, created_at_utc, global_note, pasted_text
                    FROM orders
                    WHERE customer_id = :cid AND status = 'OPEN'
                    ORDER BY id DESC
                    """
                ),
                {"cid": cid},
            ).mappings().all()

            order_ids = [int(o["order_id"]) for o in orders_rows]
            lines_rows: List[Dict] = []
            if order_ids:
                # SQLAlchemy expanding params works nicely
                lines_rows = conn.execute(
                    select(
                        order_lines.c.id.label("line_id"),
                        order_lines.c.order_id,
                        order_lines.c.card_name,
                        order_lines.c.qty,
                        order_lines.c.display_item,
                        order_lines.c.rarity,
                        order_lines.c.notes,
                        order_lines.c.raw_line,
                    )
                    .where(order_lines.c.order_id.in_(order_ids))
                    .order_by(order_lines.c.id.asc())
                ).mappings().all()

            result.append(
                {
                    "customer_id": cid,
                    "customer_name": str(c["customer_name"]),
                    "orders": [dict(o) for o in orders_rows],
                    "lines": [dict(l) for l in lines_rows],
                }
            )
        return result


def delete_open_orders_for_customer(customer_id: int) -> None:
    with ENGINE.begin() as conn:
        conn.execute(
            delete(orders).where(orders.c.customer_id == int(customer_id), orders.c.status == "OPEN")
        )


def clear_basket() -> None:
    with ENGINE.begin() as conn:
        conn.execute(delete(orders).where(orders.c.status == "OPEN"))


def update_line(line_id: int, card_name: str, qty: int, rarity: str, notes: str) -> None:
    card_name = _collapse_spaces(card_name)
    if qty < 1:
        qty = 1
    display_item = f"{card_name} {qty}"
    with ENGINE.begin() as conn:
        conn.execute(
            update(order_lines)
            .where(order_lines.c.id == int(line_id))
            .values(
                card_name=card_name,
                qty=int(qty),
                display_item=display_item,
                rarity=(rarity or ""),
                notes=(notes or ""),
            )
        )


def delete_line(line_id: int) -> None:
    with ENGINE.begin() as conn:
        conn.execute(delete(order_lines).where(order_lines.c.id == int(line_id)))


def add_manual_line_to_customer_open(customer_id: int, card_name: str, qty: int, rarity: str, notes: str) -> None:
    card_name = _collapse_spaces(card_name)
    if not card_name:
        return
    if qty < 1:
        qty = 1

    oid = create_order(int(customer_id), pasted_text="(manual)", global_note=None)
    display_item = f"{card_name} {qty}"

    with ENGINE.begin() as conn:
        conn.execute(
            insert(order_lines).values(
                order_id=int(oid),
                card_name=card_name,
                qty=int(qty),
                display_item=display_item,
                rarity=(rarity or ""),
                notes=(notes or ""),
                raw_line="(manual)",
                created_at_utc=utc_now_iso(),
            )
        )


def get_customer_open_orders_and_lines(customer_id: int) -> Tuple[str, List[Dict], List[Dict]]:
    with ENGINE.begin() as conn:
        c = conn.execute(
            select(customers.c.id, customers.c.name).where(customers.c.id == int(customer_id))
        ).mappings().first()
        if not c:
            raise ValueError("Customer not found.")

        orders_rows = conn.execute(
            sql_text(
                """
                SELECT id AS order_id, created_at_utc, global_note, pasted_text
                FROM orders
                WHERE customer_id = :cid AND status = 'OPEN'
                ORDER BY id DESC
                """
            ),
            {"cid": int(customer_id)},
        ).mappings().all()

        order_ids = [int(o["order_id"]) for o in orders_rows]
        lines_rows: List[Dict] = []
        if order_ids:
            lines_rows = conn.execute(
                select(
                    order_lines.c.id.label("line_id"),
                    order_lines.c.order_id,
                    order_lines.c.card_name,
                    order_lines.c.qty,
                    order_lines.c.display_item,
                    order_lines.c.rarity,
                    order_lines.c.notes,
                    order_lines.c.raw_line,
                )
                .where(order_lines.c.order_id.in_(order_ids))
                .order_by(order_lines.c.id.asc())
            ).mappings().all()

        return (str(c["name"]), [dict(o) for o in orders_rows], [dict(l) for l in lines_rows])


# ----------------------------
# Parsing (Hebrew + mixed English)  (UNCHANGED)
# ----------------------------
RARITY_WORDS = re.compile(r"\b(Rare|Secret|Ultra|Super|Platinum|Common|SR|UR|SCR|CR)\b", re.IGNORECASE)
HEBREW_NOTE_KEYWORDS = ["קומון", "קומונים", "פלייסט", "פלייסטים", "גרסה", "לא אכפת", "לא משנה"]

DIVIDER_RE = re.compile(r"^[\-\–\—_=]{3,}\s*$")
PAREN_RE = re.compile(r"\(([^)]*)\)")

RE_TRAIL_X = re.compile(r"^(?P<name>.+?)\s*[x×]\s*(?P<qty>\d{1,3})\s*$", re.IGNORECASE)
RE_HE_QTYWORD = re.compile(r"^(?P<name>.+?)\s+(?:כמות|עותקים|יחידות)\s*(?P<qty>\d{1,3})\s*$")
RE_TRAIL_NUM = re.compile(r"^(?P<name>.+?)\s+(?P<qty>\d{1,3})\s*$")
RE_LEAD_NUM = re.compile(r"^(?P<qty>\d{1,3})\s+(?P<name>.+?)\s*$")
RE_LEAD_NUM_STUCK = re.compile(r"^(?P<qty>\d{1,3})(?P<name>[A-Za-z].+?)\s*$")

HEBREW_RANGE = re.compile(r"[\u0590-\u05FF]")
LATIN_RANGE = re.compile(r"[A-Za-z]")


@dataclass
class ParsedLine:
    raw_line: str
    card_name_raw: str
    quantity: int
    display_item: str
    rarity: Optional[str] = None
    notes: Optional[str] = None
    needs_review: bool = False


def split_candidates(text: str) -> List[str]:
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\-_=\u2014\u2013]{3,}", "\n", text)
    raw_lines = [ln.strip() for ln in text.split("\n")]

    cleaned: List[str] = []
    for ln in raw_lines:
        if not ln:
            continue
        if DIVIDER_RE.match(ln):
            continue
        base = ln.replace(":", "").strip()
        if base in ("הזמנה", "הזמנות", "רשימה"):
            continue
        if ln.startswith("הזמנה") and not LATIN_RANGE.search(ln) and not re.search(r"\d", ln):
            continue
        cleaned.append(ln)
    return cleaned


def detect_global_note(lines: List[str]) -> Tuple[Optional[str], List[str]]:
    if not lines:
        return None, lines
    first = lines[0].strip()
    if not first:
        return None, lines
    if not HEBREW_RANGE.search(first):
        return None, lines
    if not any(k in first for k in HEBREW_NOTE_KEYWORDS):
        return None, lines
    if LATIN_RANGE.search(first) or re.search(r"\d", first):
        return None, lines
    return first, lines[1:]


def parse_line(line: str) -> ParsedLine:
    raw = line
    rarity: Optional[str] = None
    notes_parts: List[str] = []

    def _paren_repl(m: re.Match) -> str:
        nonlocal rarity
        inner = _collapse_spaces(m.group(1))
        if not inner:
            return ""
        if RARITY_WORDS.search(inner):
            rarity = inner
        else:
            notes_parts.append(inner)
        return ""

    line_wo_paren = PAREN_RE.sub(_paren_repl, line)
    line_wo_paren = _collapse_spaces(line_wo_paren)

    qty: Optional[int] = None
    name: Optional[str] = None
    for rx in (RE_TRAIL_X, RE_HE_QTYWORD, RE_TRAIL_NUM, RE_LEAD_NUM, RE_LEAD_NUM_STUCK):
        m = rx.match(line_wo_paren)
        if m:
            name = _collapse_spaces(m.group("name"))
            qty = int(m.group("qty"))
            break

    if qty is None or name is None:
        qty = 1
        name = line_wo_paren

    name = name.strip(" -•*:\t")
    name = _collapse_spaces(name)

    display = f"{name} {qty}"  # HARD RULE: qty after name
    needs_review = (not name) or (len(name) < 3) or (qty > 20)
    notes = _collapse_spaces(" | ".join(notes_parts)) if notes_parts else None

    return ParsedLine(raw, name, int(qty), display, rarity, notes, needs_review)


def parse_order(text: str) -> Tuple[Optional[str], List[ParsedLine]]:
    lines = split_candidates(text)
    global_note, remaining = detect_global_note(lines)
    parsed: List[ParsedLine] = []
    for ln in remaining:
        if "תודה" in ln and len(ln) <= 6:
            continue
        pl = parse_line(ln)
        if pl.card_name_raw:
            parsed.append(pl)
    return global_note, parsed


# ----------------------------
# Excel generation (UNCHANGED)
# ----------------------------
def style_sheet(ws, col_widths: Dict[int, float]):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

    header_font = Font(bold=True)
    header_align = Alignment(vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wrap_cols = {3, 7, 8}  # CardName, Notes, RawLine
    top_wrap = Alignment(vertical="top", wrap_text=True)
    top_nowrap = Alignment(vertical="top", wrap_text=False)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = top_wrap if cell.column in wrap_cols else top_nowrap


def build_excel_from_lines(customer_name: str, customer_id: int, lines: List[Dict], orders_: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OrderLines"
    headers = ["CustomerName", "CustomerID", "CardName", "Qty", "DisplayItem", "Rarity", "Notes", "RawLine"]
    ws.append(headers)

    order_note: Dict[int, str] = {}
    for o in orders_:
        oid = int(o["order_id"])
        gn = (o.get("global_note") or "").strip()
        order_note[oid] = gn

    for l in lines:
        oid = int(l["order_id"])
        gn = order_note.get(oid, "")
        notes = (l.get("notes") or "").strip()
        if gn:
            notes = (notes + " | " if notes else "") + gn

        ws.append(
            [
                customer_name,
                customer_id,
                l["card_name"],
                int(l["qty"]),
                l["display_item"],
                l.get("rarity") or "",
                notes,
                l.get("raw_line") or "",
            ]
        )

    style_sheet(ws, {1: 22, 2: 12, 3: 38, 4: 6, 5: 42, 6: 20, 7: 40, 8: 50})

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_excel_combined_from_db(groups: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedOrderLines"
    headers = ["CustomerName", "CustomerID", "CardName", "Qty", "DisplayItem", "Rarity", "Notes", "RawLine"]
    ws.append(headers)

    totals: Dict[str, int] = {}

    for g in groups:
        cname = g["customer_name"]
        cid = g["customer_id"]

        order_note: Dict[int, str] = {}
        for o in g["orders"]:
            order_note[int(o["order_id"])] = (o.get("global_note") or "").strip()

        for l in g["lines"]:
            oid = int(l["order_id"])
            gn = order_note.get(oid, "")
            notes = (l.get("notes") or "").strip()
            if gn:
                notes = (notes + " | " if notes else "") + gn

            ws.append(
                [
                    cname,
                    cid,
                    l["card_name"],
                    int(l["qty"]),
                    l["display_item"],
                    l.get("rarity") or "",
                    notes,
                    l.get("raw_line") or "",
                ]
            )

            key = str(l["card_name"]).strip().lower()
            totals[key] = totals.get(key, 0) + int(l["qty"])

    style_sheet(ws, {1: 22, 2: 12, 3: 38, 4: 6, 5: 42, 6: 20, 7: 40, 8: 50})

    summary = wb.create_sheet("Summary")
    summary.append(["CardName", "TotalQty"])
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = "A1:B1"
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 10
    for key in sorted(totals.keys()):
        summary.append([key, totals[key]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------
# HTML (your existing UI + ping badge)
# ----------------------------
def page_layout(body: str) -> str:
    return f"""<!doctype html>
<html lang="he">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Yugioh Orders</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 18px; }}
    .card {{ border: 1px solid #ddd; padding: 12px; border-radius: 10px; margin-top: 12px; }}
    .row {{ display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    input[type="text"] {{ padding: 8px; width: 360px; }}
    input[type="number"] {{ padding: 6px; width: 90px; }}
    textarea {{ width:100%; height: 220px; padding: 8px; }}
    button {{ padding: 8px 12px; cursor:pointer; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 10px; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; vertical-align: top; }}
    th {{ background: #f4f4f4; }}
    .muted {{ color:#666; }}
    .warn {{ background:#fff4d6; border:1px solid #f0d48a; padding:10px; border-radius:10px; }}
    .pill {{ display:inline-block; padding:2px 10px; border-radius:999px; background:#eee; margin:2px; }}
    .sectionTitle {{ display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap; }}
    .small {{ font-size: 12px; }}
    .tight td {{ padding: 6px; }}

    .pingBadge {{
      position: fixed;
      right: 14px;
      bottom: 14px;
      z-index: 9999;
      border: 1px solid #ddd;
      background: #fff;
      border-radius: 999px;
      padding: 8px 12px;
      box-shadow: 0 2px 10px rgba(0,0,0,0.08);
      font-size: 12px;
      display: flex;
      gap: 8px;
      align-items: center;
    }}
    .dot {{
      width: 10px;
      height: 10px;
      border-radius: 50%;
      background: #bbb;
      display: inline-block;
    }}
    .dot.ok {{ background: #2ecc71; }}
    .dot.bad {{ background: #e74c3c; }}
    .pingError {{
      position: fixed;
      left: 18px;
      right: 18px;
      bottom: 56px;
      z-index: 9998;
      display: none;
    }}
  </style>
</head>
<body>

<div id="pingError" class="warn pingError">
  <b>Server connection problem.</b>
  <span class="muted">Refreshing may help. If this continues, the server may be down.</span>
</div>

<div class="pingBadge" title="Auto ping every 2 minutes">
  <span id="pingDot" class="dot"></span>
  <span id="pingText" class="muted">Checking server…</span>
</div>

{body}

<script>
(function() {{
  const dot = document.getElementById("pingDot");
  const text = document.getElementById("pingText");
  const err = document.getElementById("pingError");

  function setOk(msg) {{
    dot.classList.remove("bad");
    dot.classList.add("ok");
    text.textContent = msg;
    err.style.display = "none";
  }}

  function setBad(msg) {{
    dot.classList.remove("ok");
    dot.classList.add("bad");
    text.textContent = msg;
    err.style.display = "block";
  }}

  async function doPing() {{
    const controller = new AbortController();
    const t = setTimeout(() => controller.abort(), 4000);

    try {{
      const res = await fetch("/ping", {{
        cache: "no-store",
        signal: controller.signal
      }});
      clearTimeout(t);

      if (!res.ok) {{
        setBad("Ping failed (HTTP " + res.status + ")");
        return;
      }}
      const data = await res.json();
      if (data && data.ok) {{
        setOk("Server OK");
      }} else {{
        setBad("DB/Ping failed");
      }}
    }} catch (e) {{
      clearTimeout(t);
      setBad("Server not reachable");
    }}
  }}

  doPing();
  setInterval(doPing, 120000);
}})();
</script>

</body></html>"""


# ----------------------------
# Routes (mostly unchanged)
# ----------------------------
@app.get("/", response_class=HTMLResponse)
def home(q: str = Query("", description="search customers")):
    results = search_customers(q, limit=15)
    result_pills = "".join([f'<span class="pill">{r["name"]} (ID {r["id"]})</span>' for r in results])

    groups = get_open_basket_grouped()
    basket_users = len(groups)
    basket_items = sum(len(g["lines"]) for g in groups)

    body = f"""
<h2>Paste Order → Basket</h2>

<div class="card">
  <div class="row">
    <form method="get" action="/" class="row">
      <label><b>Search users</b></label>
      <input type="text" name="q" value="{q.replace('"','&quot;')}" placeholder="Search by name (e.g. Mor)" />
      <button type="submit">Search</button>
    </form>
  </div>
  <div class="muted" style="margin-top:8px;">Results:</div>
  <div>{result_pills or "<span class='muted'>No results</span>"}</div>
</div>

<div class="card">
  <form method="post" action="/add_to_basket">
    <div class="row">
      <label><b>Recipient name (exact)</b></label>
      <input type="text" name="recipient_name" placeholder="e.g. Name Customer" required />
      <button type="submit" name="action" value="find">Find</button>
      <button type="submit" name="action" value="create">Create</button>
      <span class="muted">Use Create only if this exact name is new.</span>
    </div>

    <div style="margin-top:10px;">
      <label><b>Cards (paste WhatsApp text)</b></label>
      <textarea name="order_text" placeholder="Paste the order here..." required></textarea>
    </div>

    <div class="row" style="margin-top:10px;">
      <button type="submit" name="action" value="add">Add to Basket</button>
      <a href="/basket">Open Basket</a>
      <span class="muted">Basket: {basket_users} users, {basket_items} items</span>
    </div>
  </form>
</div>

<p class="muted">
Flow: type recipient name → click Find (must exist) or Create (new) → paste cards → Add to Basket.
</p>
"""
    return page_layout(body)


@app.post("/add_to_basket", response_class=HTMLResponse)
def add_to_basket(
    recipient_name: str = Form(...),
    order_text: str = Form(...),
    action: str = Form(...),  # find | create | add
):
    recipient_name = _collapse_spaces(recipient_name)
    order_text = order_text or ""

    if action in ("find", "create"):
        row = get_customer_by_exact_name(recipient_name)
        if action == "find":
            if not row:
                body = f"""
<h2>User not found</h2>
<div class="warn">
  "{recipient_name}" is not in the database. Use <b>Create</b> to add a new unique name.
</div>
<div style="margin-top:10px;"><a href="/">← Back</a></div>
"""
                return page_layout(body)
            body = f"""
<h2>User found</h2>
<div class="card">
  <b>{row["name"]}</b> (ID {row["id"]}) is ready. Now paste cards and press <b>Add to Basket</b>.
</div>
<div style="margin-top:10px;"><a href="/">← Back</a></div>
"""
            return page_layout(body)

        created, new_row, msg = create_customer(recipient_name)
        if new_row is None:
            body = f"""
<h2>Cannot create</h2>
<div class="warn">{msg}</div>
<div style="margin-top:10px;"><a href="/">← Back</a></div>
"""
            return page_layout(body)

        body = f"""
<h2>Create result</h2>
<div class="card">
  {msg} (ID {new_row["id"]})
  <div class="muted" style="margin-top:8px;">Now paste cards and press <b>Add to Basket</b>.</div>
</div>
<div style="margin-top:10px;"><a href="/">← Back</a></div>
"""
        return page_layout(body)

    row = get_customer_by_exact_name(recipient_name)
    if not row:
        body = f"""
<h2>Cannot add to basket</h2>
<div class="warn">
  "{recipient_name}" is not in the database.
  Click <b>Create</b> first (for a new unique name) or <b>Find</b> if it already exists.
</div>
<div style="margin-top:10px;"><a href="/">← Back</a></div>
"""
        return page_layout(body)

    customer_id = int(row["id"])
    global_note, lines = parse_order(order_text)

    order_id = create_order(customer_id, pasted_text=order_text, global_note=global_note)
    add_lines(order_id, lines)

    return RedirectResponse("/basket", status_code=303)


@app.get("/basket", response_class=HTMLResponse)
def basket():
    groups = get_open_basket_grouped()
    if not groups:
        return page_layout("""
<h2>Basket</h2>
<div class="card">
  <div class="muted">Basket is empty.</div>
  <a href="/">← Back</a>
</div>
""")

    sections_html = ""
    for g in groups:
        cid = g["customer_id"]
        cname = g["customer_name"]

        rows_html = ""
        for l in g["lines"]:
            line_id = int(l["line_id"])
            rows_html += f"""
<tr>
  <td class="small muted">#{line_id}</td>
  <td>
    <form method="post" action="/line_update" class="row" style="gap:6px;">
      <input type="hidden" name="line_id" value="{line_id}"/>
      <input type="hidden" name="customer_id" value="{cid}"/>
      <input type="text" name="card_name" value="{str(l["card_name"]).replace('"','&quot;')}" style="width:280px;" required/>
  </td>
  <td>
      <input type="number" name="qty" value="{int(l["qty"])}" min="1" />
  </td>
  <td>
      <input type="text" name="rarity" value="{str(l.get("rarity") or '').replace('"','&quot;')}" style="width:160px;" />
  </td>
  <td>
      <input type="text" name="notes" value="{str(l.get("notes") or '').replace('"','&quot;')}" style="width:260px;" />
  </td>
  <td>
      <button type="submit">Save</button>
    </form>
  </td>
  <td>
    <form method="post" action="/line_delete">
      <input type="hidden" name="line_id" value="{line_id}"/>
      <button type="submit">Delete</button>
    </form>
  </td>
</tr>
"""

        add_row = f"""
<tr>
  <td class="small muted">+</td>
  <td colspan="6">
    <form method="post" action="/line_add" class="row" style="gap:6px;">
      <input type="hidden" name="customer_id" value="{cid}"/>
      <input type="text" name="card_name" placeholder="New card name" style="width:280px;" required/>
      <input type="number" name="qty" value="1" min="1" />
      <input type="text" name="rarity" placeholder="Rarity (optional)" style="width:160px;" />
      <input type="text" name="notes" placeholder="Notes (optional)" style="width:260px;" />
      <button type="submit">Add</button>
    </form>
  </td>
</tr>
"""

        sections_html += f"""
<div class="card">
  <div class="sectionTitle">
    <div>
      <b>{cname}</b> <span class="muted">(ID {cid})</span>
      <div class="muted small">Lines: {len(g["lines"])} | Orders (messages): {len(g["orders"])}</div>
    </div>
    <div class="row">
      <form method="post" action="/download_user">
        <input type="hidden" name="customer_id" value="{cid}"/>
        <button type="submit">Download Excel</button>
      </form>
      <form method="post" action="/remove_user">
        <input type="hidden" name="customer_id" value="{cid}"/>
        <button type="submit">Remove User From Basket</button>
      </form>
    </div>
  </div>

  <table class="tight">
    <thead>
      <tr>
        <th>#</th>
        <th>CardName</th>
        <th>Qty</th>
        <th>Rarity</th>
        <th>Notes</th>
        <th>Save</th>
        <th>Delete</th>
      </tr>
    </thead>
    <tbody>
      {rows_html or "<tr><td colspan='7' class='muted'>No items.</td></tr>"}
      {add_row}
    </tbody>
  </table>
</div>
"""

    basket_users = len(groups)
    basket_items = sum(len(g["lines"]) for g in groups)

    body = f"""
<h2>Basket</h2>

<div class="card">
  <div class="row">
    <a href="/">← Back</a>
    <span class="muted">Users: {basket_users} | Items: {basket_items}</span>
    <form method="post" action="/download_combined">
      <button type="submit">Download Combined Excel</button>
    </form>
    <form method="post" action="/clear_basket">
      <button type="submit">Clear Basket</button>
    </form>
  </div>
</div>

{sections_html}
"""
    return page_layout(body)


@app.post("/line_update")
def line_update(
    line_id: int = Form(...),
    customer_id: int = Form(...),
    card_name: str = Form(...),
    qty: int = Form(...),
    rarity: str = Form(""),
    notes: str = Form(""),
):
    update_line(line_id=int(line_id), card_name=card_name, qty=int(qty), rarity=rarity or "", notes=notes or "")
    return RedirectResponse("/basket", status_code=303)


@app.post("/line_delete")
def line_delete(line_id: int = Form(...)):
    delete_line(int(line_id))
    return RedirectResponse("/basket", status_code=303)


@app.post("/line_add")
def line_add(
    customer_id: int = Form(...),
    card_name: str = Form(...),
    qty: int = Form(...),
    rarity: str = Form(""),
    notes: str = Form(""),
):
    add_manual_line_to_customer_open(int(customer_id), card_name, int(qty), rarity or "", notes or "")
    return RedirectResponse("/basket", status_code=303)


@app.post("/remove_user")
def remove_user(customer_id: int = Form(...)):
    delete_open_orders_for_customer(int(customer_id))
    return RedirectResponse("/basket", status_code=303)


@app.post("/clear_basket")
def clear_basket_route():
    clear_basket()
    return RedirectResponse("/basket", status_code=303)


@app.post("/download_user")
def download_user(customer_id: int = Form(...)):
    cid = int(customer_id)
    cname, orders_, lines = get_customer_open_orders_and_lines(cid)

    if not lines:
        return RedirectResponse("/basket", status_code=303)

    data = build_excel_from_lines(cname, cid, lines, orders_)
    today = datetime.now().date().isoformat()
    safe_name = _collapse_spaces(cname).replace("/", "-").replace("\\", "-")
    filename = f"{today} - {safe_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/download_combined")
def download_combined():
    groups = get_open_basket_grouped()
    if not groups:
        return RedirectResponse("/basket", status_code=303)

    data = build_excel_combined_from_db(groups)
    today = datetime.now().date().isoformat()
    filename = f"{today} - Combined Orders.xlsx"

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/healthz")
def healthz(response: Response):
    t0 = time.perf_counter()
    try:
        conn = db_connect()
        try:
            conn.execute("SELECT 1;")
            # Optional: verify the tables exist
            conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='customers';")
        finally:
            conn.close()
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "db": "ok", "latency_ms": ms, "utc": utc_now_iso()}
    except Exception as e:
        response.status_code = 503
        return {"ok": False, "db": "fail", "error": str(e), "utc": utc_now_iso()}